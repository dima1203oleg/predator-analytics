#!/usr/bin/env python3
"""
Простий скрипт для підрахунку рядків в Excel файлі
Використовується автономним E2E тестом для перевірки консистентності
"""

import sys
import json
import openpyxl

def count_excel_rows(file_path: str) -> dict:
    """
    Підраховує кількість рядків в Excel файлі
    
    Args:
        file_path: Шлях до Excel файлу (.xlsx)
    
    Returns:
        dict: JSON з інформацією про файл та кількість рядків
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        total_rows = 0
        sheet_info = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            row_count = sheet.max_row - 1  # Віднімаємо 1 для заголовка
            
            if row_count > 0:
                total_rows += row_count
                sheet_info.append({
                    "sheet_name": sheet_name,
                    "row_count": row_count,
                    "max_column": sheet.max_column
                })
        
        workbook.close()
        
        return {
            "success": True,
            "file_path": file_path,
            "total_rows": total_rows,
            "sheets": sheet_info,
            "sheet_count": len(sheet_info)
        }
        
    except FileNotFoundError:
        return {
            "success": False,
            "error": "File not found",
            "file_path": file_path
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "file_path": file_path
        }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "Usage: python excel_row_counter.py <file_path>"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    result = count_excel_rows(file_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    sys.exit(0 if result["success"] else 1)