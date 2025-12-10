"""
Test Data Generator

Generates realistic customs declarations test data for E2E testing.
Creates XLSX files with configurable row count.
"""

import os
import random
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger("service.test_data_generator")

# Constants for realistic data
COUNTRIES = [
    ("CN", "Китай"), ("DE", "Німеччина"), ("US", "США"), ("PL", "Польща"),
    ("TR", "Туреччина"), ("IT", "Італія"), ("FR", "Франція"), ("CZ", "Чехія"),
    ("IN", "Індія"), ("KR", "Республіка Корея"), ("JP", "Японія"), ("GB", "Великобританія"),
    ("NL", "Нідерланди"), ("BE", "Бельгія"), ("ES", "Іспанія")
]

CUSTOMS_OFFICES = [
    "Київська митниця", "Одеська митниця", "Львівська митниця",
    "Харківська митниця", "Дніпровська митниця", "Чернігівська митниця",
    "Волинська митниця", "Закарпатська митниця"
]

# HS Codes with descriptions (realistic examples)
HS_CODES = [
    ("8471.30", "Комп'ютерне обладнання портативне"),
    ("8542.31", "Мікросхеми інтегральні цифрові"),
    ("6110.20", "Одяг трикотажний бавовняний"),
    ("8703.23", "Автомобілі легкові бензинові"),
    ("3004.10", "Лікарські засоби дозовані"),
    ("8517.12", "Телефони мобільні смартфони"),
    ("4202.21", "Сумки шкіряні"),
    ("9401.30", "Меблі для сидіння офісні"),
    ("8528.51", "Монітори LED дисплеї"),
    ("2204.21", "Вино виноградне натуральне"),
    ("8418.10", "Холодильники побутові"),
    ("8443.32", "Принтери та БФП"),
    ("8471.41", "Комп'ютери системні блоки"),
    ("8504.40", "Перетворювачі електричні"),
    ("8544.42", "Кабелі електричні"),
    ("9403.20", "Меблі металеві"),
    ("6204.62", "Штани жіночі бавовняні"),
    ("6203.42", "Штани чоловічі бавовняні"),
    ("8507.60", "Акумулятори літієві"),
    ("8523.51", "Накопичувачі SSD")
]

COMPANY_PREFIXES = [
    "ТОВ", "ПП", "ПрАТ", "АТ", "ФОП"
]

COMPANY_NAMES = [
    "Техноімпорт", "ЄвроТрейд", "Глобал Логістик", "Укрторг",
    "СхідЗахід", "Нова Лінія", "Альфа Груп", "Омега Трейдинг",
    "Прайм Інвест", "Сіті Комерс", "Юнітрейд", "Макс Імпорт"
]


class TestDataGenerator:
    """Generates realistic test data for customs declarations"""
    
    def __init__(self):
        self.generated_declaration_numbers = set()
    
    def generate_edrpou(self) -> str:
        """Generate realistic Ukrainian company registration number"""
        return str(random.randint(10000000, 99999999))
    
    def generate_declaration_number(self, year: int = 2024) -> str:
        """Generate unique declaration number"""
        while True:
            number = f"UA{random.randint(100000, 999999)}/{year % 100}"
            if number not in self.generated_declaration_numbers:
                self.generated_declaration_numbers.add(number)
                return number
    
    def generate_company_name(self) -> str:
        """Generate realistic Ukrainian company name"""
        prefix = random.choice(COMPANY_PREFIXES)
        name = random.choice(COMPANY_NAMES)
        suffix = random.choice(["", " Україна", " Груп", " Плюс", ""])
        return f"{prefix} «{name}{suffix}»"
    
    def generate_declaration(self, date: datetime) -> Dict[str, Any]:
        """Generate a single customs declaration record"""
        country_code, country_name = random.choice(COUNTRIES)
        hs_code, description = random.choice(HS_CODES)
        
        weight = round(random.uniform(0.5, 5000), 2)
        value = round(random.uniform(100, 500000), 2)
        customs_office = random.choice(CUSTOMS_OFFICES)
        
        # 95% chance of successful processing
        status = "Оформлено" if random.random() > 0.05 else random.choice([
            "В обробці", "Потребує перевірки", "Затримано"
        ])
        
        return {
            "Номер декларації": self.generate_declaration_number(date.year),
            "Дата": date.strftime("%d.%m.%Y"),
            "Код товару (HS)": hs_code,
            "Опис товару": description,
            "Країна походження (код)": country_code,
            "Країна походження": country_name,
            "Митна вартість (USD)": value,
            "Вага брутто (кг)": weight,
            "Митниця оформлення": customs_office,
            "Отримувач (ЄДРПОУ)": self.generate_edrpou(),
            "Отримувач (назва)": self.generate_company_name(),
            "Статус": status,
            "Мито (USD)": round(value * 0.1, 2),
            "ПДВ (USD)": round(value * 0.2, 2),
            "Кількість": random.randint(1, 1000),
            "Одиниця виміру": random.choice(["шт.", "кг", "л", "м", "комплект"])
        }
    
    def generate_batch(
        self, 
        row_count: int = 500, 
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> List[Dict[str, Any]]:
        """Generate a batch of customs declarations"""
        if start_date is None:
            start_date = datetime(2024, 3, 1)
        if end_date is None:
            end_date = datetime(2024, 3, 31)
        
        date_range = (end_date - start_date).days
        
        declarations = []
        for _ in range(row_count):
            random_days = random.randint(0, date_range)
            declaration_date = start_date + timedelta(days=random_days)
            declarations.append(self.generate_declaration(declaration_date))
        
        # Sort by date
        declarations.sort(key=lambda x: datetime.strptime(x["Дата"], "%d.%m.%Y"))
        
        return declarations
    
    def generate_xlsx(
        self, 
        output_path: str, 
        row_count: int = 500,
        sheet_name: str = "Митні декларації"
    ) -> Dict[str, Any]:
        """Generate XLSX file with test data"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Using basic xlsx generation.")
            return self._generate_basic_xlsx(output_path, row_count)
        
        # Generate data
        data = self.generate_batch(row_count)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Headers
        headers = list(data[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Data rows
        for row_idx, record in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record[header])
                cell.border = thin_border
                
                # Format currency columns
                if "USD" in header:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ""))
                for row in range(1, min(row_count + 2, 100))
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        
        logger.info(f"Generated {row_count} records to {output_path}")
        
        return {
            "success": True,
            "path": output_path,
            "row_count": row_count,
            "columns": headers,
            "file_size": os.path.getsize(output_path)
        }
    
    def _generate_basic_xlsx(self, output_path: str, row_count: int) -> Dict[str, Any]:
        """Fallback basic XLSX generation without openpyxl"""
        try:
            import xlsxwriter
            
            data = self.generate_batch(row_count)
            headers = list(data[0].keys())
            
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            
            workbook = xlsxwriter.Workbook(output_path)
            worksheet = workbook.add_worksheet("Митні декларації")
            
            # Header format
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'align': 'center'
            })
            
            # Write headers
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
            
            # Write data
            for row_idx, record in enumerate(data, 1):
                for col_idx, header in enumerate(headers):
                    worksheet.write(row_idx, col_idx, record[header])
            
            workbook.close()
            
            return {
                "success": True,
                "path": output_path,
                "row_count": row_count,
                "columns": headers,
                "file_size": os.path.getsize(output_path)
            }
            
        except ImportError:
            logger.error("Neither openpyxl nor xlsxwriter available")
            return {
                "success": False,
                "error": "No XLSX library available"
            }


# Singleton
_generator = TestDataGenerator()

def get_test_data_generator() -> TestDataGenerator:
    return _generator


# CLI usage
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate test customs data")
    parser.add_argument("--rows", type=int, default=500, help="Number of rows to generate")
    parser.add_argument("--output", type=str, default="sample_data/Березень_2024.xlsx", help="Output file path")
    
    args = parser.parse_args()
    
    generator = TestDataGenerator()
    result = generator.generate_xlsx(args.output, args.rows)
    
    if result["success"]:
        print(f"✅ Generated {result['row_count']} records to {result['path']}")
        print(f"   File size: {result['file_size'] / 1024:.1f} KB")
    else:
        print(f"❌ Failed: {result.get('error')}")
